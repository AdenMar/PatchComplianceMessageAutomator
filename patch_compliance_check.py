# all data should be in sheet titled all within workbook
# data should be organized by site
# status indicators should be "on" and "off" - not anymore

import openpyxl
import sys
import datetime

# args to add:
# filter (on/off)
# customizable open and closer

def main(path, saveDir):
    wb = openpyxl.load_workbook(path[1])
    sh = wb["all"]
    outF = open(saveDir, "a")

    header = []

    for col in range(1,sh.max_column+1):
        # print(sh.cell(1,col).value)
        header.append(sh.cell(1,col).value)
    
    colSite = header.index("Site Name") + 1
    colLastUser = header.index("Last User") + 1
    colName = header.index("Name") + 1
    colStatus = header.index("Availability") + 1

    siteHolder = sh.cell(2,colSite).value
    newMessage(siteHolder,outF)

    # itterate through rows (start at 2 avoid header)
    # index bySite [n][0] is on [n][1] is off
    for row in range(2, sh.max_row+2): 
        currSite = sh.cell(row,colSite).value
        currLastUser = sh.cell(row,colLastUser).value
        currName = sh.cell(row,colName).value
        currStatus = sh.cell(row,colStatus).value

        if (currSite != siteHolder):
            newMessage(currSite,outF)
            # return to defaults
            siteHolder = currSite

        writeAndLog("Computer Name: " + str(currName),outF)
        writeAndLog("Last Logged in User: " + str(currLastUser),outF)

        if (currStatus == "true"):
            writeAndLog("Issue: Your PC is out of date",outF)
            writeAndLog("Action to be taken: Schedule a time with us for a reboot",outF)
        elif (currStatus == "false"):
            writeAndLog("Issue: Your PC is offline",outF)
            writeAndLog("Action to be taken: Power on the system or let us know if it is no longer in use",outF)

        writeAndLog("",outF)

def writeAndLog(content, file):
    file.write(content + "\n")
    # for debug perposes
    # print("File Updated")
    # print(content)

def newMessage(site, file):
    writeAndLog("_____" + str(site) + "_____\n",file)
    writeAndLog("We proactively monitor your computers for missing Critical and Security Updates of the Operating System and have identified the following issue(s) that require attention.\n",file)
    writeAndLog("Please review and let us know if you would like assistance with the action to be taken.\n\n",file)

if __name__=="__main__":
    date = datetime.datetime.utcnow()
    newFile = "GetWell_" + str(date).replace(":","-") + ".txt"
    main(sys.argv, newFile)